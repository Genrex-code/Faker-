"""
NAHUAL - Export Manager (exports/ExportManager.py)
Responsabilidad Única: Gestionar TODAS las exportaciones a diferentes formatos.

Principios:
- Un solo punto de entrada: exportar()
- Cada formato tiene su propia función privada
- Pipilinais SOLO llama a este módulo
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import pandas as pd

logger = logging.getLogger(__name__)


def exportar(
    dataset: Dict[str, List[Any]],
    formato: str,
    volumen: int,
    metadata: Optional[Dict[str, Any]] = None
) -> bool:
    """
    Punto único de entrada para exportar datasets.
    
    Args:
        dataset: Diccionario columna -> lista de valores
        formato: 'excel', 'csv', 'json'
        volumen: Número de registros
        metadata: Info adicional (tiempo_generacion, motor, etc.)
    
    Returns:
        True si éxito, False si error
    """
    logger.info(f"🚚 [EXPORT MANAGER] Exportando a {formato.upper()}...")
    
    # Validaciones
    if not dataset or len(dataset) == 0:
        logger.error("❌ Dataset vacío")
        return False
    
    # Mapeo de formatos
    exportadores = {
        'excel': _exportar_excel,
        'csv': _exportar_csv,
        'json': _exportar_json,
    }
    
    exportador = exportadores.get(formato.lower())
    
    if not exportador:
        logger.error(f"❌ Formato '{formato}' no soportado")
        logger.info(f"   Formatos disponibles: {list(exportadores.keys())}")
        return False
    
    # Ejecutar exportador específico
    try:
        return exportador(dataset, volumen, metadata)
    except Exception as e:
        logger.error(f"💥 Error en exportación {formato}: {e}", exc_info=True)
        return False


# ============================================================================
# EXPORTADORES ESPECÍFICOS (PRIVADOS)
# ============================================================================

def _exportar_excel(
    dataset: Dict[str, List[Any]],
    volumen: int,
    metadata: Optional[Dict[str, Any]] = None
) -> bool:
    """Exporta a Excel con estilos profesionales"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear directorio
        output_dir = Path("outputs")
        output_dir.mkdir(exist_ok=True)
        
        # Nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = output_dir / f"nahual_datos_{timestamp}.xlsx"
        
        # Convertir a DataFrame
        df = pd.DataFrame(dataset)
        
        # Escribir Excel con estilos
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # ESTILOS
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Aplicar estilos a encabezados
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Autoajustar columnas
            for col in range(1, len(df.columns) + 1):
                max_len = 0
                col_letter = get_column_letter(col)
                for row in range(1, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = min(max_len + 3, 50)
            
            # Congelar panel
            worksheet.freeze_panes = 'A2'
            
            # Filtros automáticos
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Hoja de metadata si existe
        if metadata:
            _agregar_hoja_metadata(workbook, metadata, volumen, len(df.columns))
        
        logger.info(f"✅ Excel guardado: {ruta}")
        logger.info(f"   📊 {volumen:,} filas × {len(dataset)} columnas")
        return True
        
    except ImportError as e:
        logger.error(f"❌ Dependencia faltante: {e}")
        logger.info("   Instala: pip install pandas openpyxl")
        return False
    except Exception as e:
        logger.error(f"❌ Error exportando a Excel: {e}")
        return False


def _exportar_csv(
    dataset: Dict[str, List[Any]],
    volumen: int,
    metadata: Optional[Dict[str, Any]] = None
) -> bool:
    """Exporta a CSV (formato universal)"""
    try:
        output_dir = Path("outputs")
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = output_dir / f"nahual_datos_{timestamp}.csv"
        
        df = pd.DataFrame(dataset)
        df.to_csv(ruta, index=False, encoding='utf-8-sig')
        
        logger.info(f"✅ CSV guardado: {ruta}")
        logger.info(f"   📊 {volumen:,} filas × {len(dataset)} columnas")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error exportando a CSV: {e}")
        return False


def _exportar_json(
    dataset: Dict[str, List[Any]],
    volumen: int,
    metadata: Optional[Dict[str, Any]] = None
) -> bool:
    """Exporta a JSON (formato para APIs)"""
    try:
        import json
        
        output_dir = Path("outputs")
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = output_dir / f"nahual_datos_{timestamp}.json"
        
        # Convertir a lista de registros
        df = pd.DataFrame(dataset)
        registros = df.to_dict(orient='records')
        
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(registros, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"✅ JSON guardado: {ruta}")
        logger.info(f"   📊 {volumen:,} registros")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error exportando a JSON: {e}")
        return False


def _agregar_hoja_metadata(workbook, metadata: Dict[str, Any], volumen: int, columnas: int):
    """Agrega hoja con metadatos de generación (solo Excel)"""
    try:
        if 'Metadata' in workbook.sheetnames:
            return
        
        ws = workbook.create_sheet('Metadata')
        
        # Información de generación
        ws['A1'] = 'NAHUAL - Metadatos de Generación'
        ws['A1'].font = Font(bold=True, size=12)
        
        metadatos = [
            ('Fecha generación', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('Volumen', f"{volumen:,} registros"),
            ('Columnas', columnas),
            ('Motor usado', metadata.get('motor', 'N/A')),
            ('Tiempo generación', f"{metadata.get('tiempo_generacion', 0):.2f} segundos"),
        ]
        
        for i, (key, value) in enumerate(metadatos, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Autoajustar
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
    except Exception as e:
        logger.debug(f"No se pudo agregar metadata: {e}")


# ============================================================================
# PRUEBA DEL EXPORT MANAGER
# ============================================================================
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    
    print("\n🧪 Probando Export Manager...")
    
    # Datos de prueba
    test_data = {
        'nombre': ['Ana', 'Juan', 'Maria'],
        'edad': [25, 30, 28],
        'ciudad': ['CDMX', 'GDL', 'MTY']
    }
    
    # Probar Excel
    exportar(test_data, 'excel', 3, {'motor': 'test', 'tiempo_generacion': 0.5})
    
    # Probar CSV
    exportar(test_data, 'csv', 3, None)
    
    # Probar JSON
    exportar(test_data, 'json', 3, None)
    
    print("\n✅ Pruebas completadas. Revisa la carpeta 'outputs/'")