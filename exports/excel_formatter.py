"""
NAHUAL - Excel Formatter (exports/excel_formatter.py)
Responsabilidad: Exportar a Excel con estilos profesionales
"""

import logging
from pathlib import Path
from datetime import datetime
import pandas as pd

logger = logging.getLogger(__name__)


def exportar_excel(dataset: dict, volumen: int, ruta: str = None, metadata: dict = None) -> bool:
    """
    Exporta dataset a Excel con formato profesional
    
    Args:
        dataset: Diccionario con los datos
        volumen: Número de registros
        ruta: Ruta de destino (opcional)
        metadata: Metadatos para hoja adicional
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear directorio si no existe
        output_dir = Path("outputs")
        output_dir.mkdir(exist_ok=True)
        
        # Generar nombre de archivo si no se proporcionó
        if not ruta:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = output_dir / f"nahual_datos_{timestamp}.xlsx"
        else:
            ruta = Path(ruta)
        
        # Convertir a DataFrame
        df = pd.DataFrame(dataset)
        
        # Escribir Excel con estilos
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Estilos profesionales
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Aplicar estilos a encabezados
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                worksheet.row_dimensions[1].height = 25
            
            # Autoajustar columnas
            for col in range(1, len(df.columns) + 1):
                max_len = 0
                col_letter = get_column_letter(col)
                for row in range(1, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = min(max_len + 3, 50)
            
            # Congelar panel y filtros
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Agregar hoja de metadata si existe
            if metadata:
                _agregar_metadata_excel(workbook, metadata, volumen, len(df.columns))
        
        logger.info(f"✅ Excel guardado: {ruta}")
        logger.info(f"   📊 {volumen:,} filas × {len(dataset)} columnas")
        return True
        
    except ImportError as e:
        logger.error(f"❌ Dependencia faltante: {e}")
        logger.info("   Instala: pip install pandas openpyxl")
        return False
    except Exception as e:
        logger.error(f"❌ Error exportando a Excel: {e}")
        return False


def _agregar_metadata_excel(workbook, metadata: dict, volumen: int, columnas: int):
    """Agrega hoja de metadatos al Excel"""
    try:
        from openpyxl.styles import Font
        
        if 'Metadata' in workbook.sheetnames:
            return
        
        ws = workbook.create_sheet('Metadata')
        
        ws['A1'] = 'NAHUAL - Metadatos de Generación'
        ws['A1'].font = Font(bold=True, size=12)
        
        metadatos = [
            ('Fecha generación', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('Volumen', f"{volumen:,} registros"),
            ('Columnas', columnas),
            ('Motor usado', metadata.get('motor', 'N/A')),
            ('Tiempo generación', f"{metadata.get('tiempo_generacion', 0):.2f} segundos"),
        ]
        
        for i, (key, value) in enumerate(metadatos, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        
    except Exception as e:
        logger.debug(f"No se pudo agregar metadata: {e}")